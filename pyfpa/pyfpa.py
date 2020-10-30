"""
Project for Financial Planning and Analysis

Written by Erik Warren
Written for Rebecca, Ethan and Clare
October 2020
version: 0.0.1 beta
"""

import pandas as pd
import numpy as np
import os
import datetime as dt
from openpyxl import load_workbook


class fpa:
    fpa_help = "pyfpa - Financial Planning and Analysis Python Project.  Python intro for FP&A Professionals."

    def __init__(self, df=pd.DataFrame()):
        self.accounts = pd.DataFrame()  # Repository for Chart of Accounts or other
        # lists to add to dimensions
        self.base = df.copy()  # Original import data for block
        self.block = df.copy()  # Imported data block
        self.meta_block = pd.DataFrame()   # Meta data for block.  Filepath, last access/mod
        self.data = pd.DataFrame()  # The main dataframe for all the data
        self.consolidation = pd.DataFrame()  # Container for consolidate() result
        self.slice = pd.DataFrame()  # Container for slicing and dicing .data
        self.variance = pd.DataFrame()  # General container for functions
        self.function_result= []  # General container for functions
        # TODO: Add Resample
        # Index names can only be strings

    help_import_xl = '''Import Excel or DataFrame to .block.  cols_to_index are columns for index \
    which can be int or list for Excel. For DataFrame, string or list of string column names.'''

    def import_xl(self, fpath, ws_name=0, cols_to_index=0):
        """
        Import a table from a worksheet in a Excel File or an existing pandas DataFrame.

        :param fpath: path to file OR a pandas DataFrame
        :param ws_name: worksheet name or index such as 0 or 2
        :param cols_to_index: columns to put into index either a number or a list i.e. [0, 1, 2]
        """
        if isinstance(fpath, pd.DataFrame):
            if cols_to_index != 0:
                fpath.set_index(cols_to_index, append=True, inplace=True)
            self.block = fpath.copy()
        else:
            self.block = pd.read_excel(
                fpath, sheet_name=ws_name, index_col=cols_to_index
            )
        self.block.dropna(0, "all", inplace=True)
        self.block.dropna(1, "all", inplace=True)
        self.base = self.block.copy()
        data_block_value = 0
        new_index_names = []
        if not isinstance(self.block.index, pd.MultiIndex):
            new_index_names = ["Data_Block", self.block.index.name]
        if not isinstance(self.data.index, pd.MultiIndex):
            if not isinstance(self.block.index, pd.MultiIndex):
                data_block_value = int(np.random.rand() * 1000000000)
                new_index_values = [[data_block_value], list(self.block.index)]
                self.block.index = pd.MultiIndex.from_product(
                    new_index_values, names=new_index_names
                )
            else:
                data_block_value = int(np.random.rand() * 1000000000)
                dfi = self.block.index.to_frame()
                dfi.insert(1, "Data_Block", data_block_value)
                self.block.index = pd.MultiIndex.from_frame(dfi)
        else:
            data_block_no_dup = pd.Series(
                self.data.index.get_level_values("Data_Block")
            )
            while data_block_value in data_block_no_dup:
                data_block_value = int(np.random.rand() * 1000000000)
            dfi = self.block.index.to_frame()
            dfi.insert(1, "Data_Block", data_block_value)
            self.block.index = pd.MultiIndex.from_frame(dfi)
        # Meta_df
        meta_df = pd.DataFrame(
            index=[data_block_value],
            columns=[
                "file",
                "last_modified",
                "last_accessed",
                "file_path",
                "modified_by",
            ],
        )
        file_stats = os.stat(fpath) if isinstance(fpath, str) else "calculated block"
        meta_df.loc[data_block_value, "file"] = (
            fpath.split("\\")[-1] if isinstance(fpath, str) else file_stats
        )
        meta_df.loc[data_block_value, "file_path"] = (
            fpath if isinstance(fpath, str) else file_stats
        )
        meta_df.loc[data_block_value, "last_modified"] = (
            dt.datetime.fromtimestamp(file_stats.st_mtime).strftime("%Y-%m-%d %H:%M")
            if isinstance(fpath, str)
            else dt.datetime.now().strftime("%Y-%m-%d %H:%M")
        )
        meta_df.loc[data_block_value, "last_accessed"] = (
            dt.datetime.fromtimestamp(file_stats.st_atime).strftime("%Y-%m-%d %H:%M")
            if isinstance(fpath, str)
            else dt.datetime.now().strftime("%Y-%m-%d %H:%M")
        )
        self.meta_block = self.meta_block.append(meta_df)

    help_import_custom_xl = '''Import Excel using custom mapping for table and dimensions.  Enter most items as lists.'''

    def import_custom_xl(
        self,
        f_path,
        ws_name=0,
        table_coords=None,
        idx_cols=0,
        dim_names=None,
        dim_values=None,
        dim_names_coords=None,
        dim_coords=None,
        fill_index_na=False,
    ):
        """
        Custom import with mapping for data table and dimensions from Excel worksheet.

        :param f_path: Path to the Excel File.
        :param ws_name: Name or index number of the worksheet.  i.e. 'Accounts' or 2.
        :param table_coords: Excel references for top left and bottom right of table.  i.e. ['B7', 'Z20'].
        :param idx_cols: Number of index columns to use.  0 would be the first column and 3 would mean first 3 columns.
        :param dim_names: New dimensions names to add.  i.e. ['Department', 'Geography']
        :param dim_values: Values for new dimensions.  i.e. ['Sales', 'North America']
        :param dim_names_coords: Excel references to get dimension names from the Excel.  i.e. ['A3', 'A4'].
        :param dim_coords: Excel references to get dimension values from the Excel.  i.e. ['B3', 'B4'].
        """
        self.block = pd.DataFrame()
        wb = load_workbook(f_path, data_only=True)
        ws_name = wb.sheetnames[0] if ws_name == 0 else ws_name
        ws = wb[ws_name]
        dim_names = [] if dim_names is None else dim_names
        if dim_names_coords.__ne__(None) and dim_names_coords.__ne__([None]):
            for item in dim_names_coords:
                dim_name = ws[item].value
                dim_names.append(dim_name)
        dim_values = [] if dim_values is None else dim_values
        if dim_coords.__ne__(None) and dim_coords.__ne__([None]):
            for item in dim_coords:
                dim_value = ws[item].value
                dim_values.append(dim_value)
        if table_coords == None or table_coords == [None]:
            df = pd.read_excel(f_path, sheet_name=ws_name, index_col=idx_cols)
        else:
            data_range = ws[table_coords[0] : table_coords[1]]
            range_cols, range_index, range_values, row_list = [], [], [], []
            idx_name = data_range[0][0]
            for item in data_range[0][1:]:
                range_cols.append(item.value)
            col_width = len(range_cols)
            for item in data_range[1:]:
                range_index.append(item[0].value)
            for item in data_range[1:]:
                for itemx in item[1:]:
                    range_values.append(itemx.value)
            row_count = int(len(range_values) / col_width)
            col_width_start, col_width_end = 0, col_width
            for item in list(range(1, row_count + 1)):
                append_list = range_values[col_width_start:col_width_end]
                row_list.append(append_list)
                col_width_start += col_width
                col_width_end += col_width
            df = pd.DataFrame(row_list, index=range_index, columns=range_cols)
            df.index.name = idx_name.value
            if idx_cols.__ne__(None):
                add_idx_range = list(range(0, idx_cols - 1))
                for item in add_idx_range:
                    add_col_name = range_cols[item]
                    df.set_index(add_col_name, append=True, inplace=True)
        df.dropna(0, "all", inplace=True)
        df.dropna(1, "all", inplace=True)

        self.block = df
        self.base = self.block.copy()
        data_block_value = 0
        if not isinstance(self.data.index, pd.MultiIndex):
            data_block_value = int(np.random.rand() * 1000000000)
            dfi = self.block.index.to_frame()
            dfi["Data_Block"] = data_block_value
            rev_cols = list(dfi.columns)
            rev_cols.reverse()
            dfi = dfi.loc[:, rev_cols]
            if fill_index_na == True:
                dfi.fillna(method="ffill", inplace=True)
            self.block.index = pd.MultiIndex.from_frame(dfi)
        else:
            data_block_no_dup = pd.Series(
                self.data.index.get_level_values("Data_Block")
            )
            while data_block_value in data_block_no_dup:
                data_block_value = int(np.random.rand() * 1000000000)
            dfi = self.block.index.to_frame()
            dfi["Data_Block"] = data_block_value
            rev_cols = list(dfi.columns)
            rev_cols.reverse()
            dfi = dfi.loc[:, rev_cols]
            if fill_index_na == True:
                dfi.fillna(method="ffill", inplace=True)
            self.block.index = pd.MultiIndex.from_frame(dfi)
        if dim_values.__ne__([None]) and dim_names.__ne__([None]):
            self.add_dimensions(dim_names, dim_values, data_obj='block')
        #Add meta data
        meta_df = pd.DataFrame(
            index=[data_block_value],
            columns=[
                "file",
                "import_time",
                "last_modified",
                "last_accessed",
                "file_path",
                "modified_by",
            ],
        )
        file_stats = os.stat(f_path)
        meta_df.loc[data_block_value, "file"] = (
            f_path.split("\\")[-1] if f_path.count("\\") > 0 else f_path.split("/")[-1]
        )
        meta_df.loc[data_block_value, "file_path"] = f_path
        meta_df.loc[data_block_value, "import_time"] = dt.datetime.now()
        meta_df.loc[data_block_value, "last_modified"] = dt.datetime.fromtimestamp(
            file_stats.st_mtime
        ).strftime("%Y-%m-%d %H:%M")
        meta_df.loc[data_block_value, "last_accessed"] = dt.datetime.fromtimestamp(
            file_stats.st_atime
        ).strftime("%Y-%m-%d %H:%M")
        self.meta_block = self.meta_block.append(meta_df)

    def import_accts_xl(self, f_path, ws_name=0, dim_name='nval'):
        """
        Import a dataframe such as a chart of accounts or sales dimensions for adding to data objects.

        :param f_path: Path to Excel file.
        :param ws_name: Name or index number of the worksheet.  i.e. 'Accounts' or 2.
        :param dim_name: Identifier for the group used when retrieving it.  See merge_dim_from_accts.
        :return: self.accounts
        """
        dim_set = pd.read_excel(f_path, ws_name)
        dim_set.index.name = "index"
        dim_set["dim_set"] = dim_name
        dim_set.set_index("dim_set", append=True, inplace=True)
        dim_set.index = dim_set.index.reorder_levels(["dim_set", "index"])
        self.accounts = pd.concat([self.accounts, dim_set])
        # TODO: Add check to see if adding an identical df. Concat with different indices.

    def import_xl_sheets(
        self,
        f_path,
        wb_sheets=None,
        table_coords=None,
        idx_cols=0,
        dim_names=None,
        dim_vals=None,
        dim_names_coords=None,
        dim_coords=None,
        fill_index_na=False,
    ):
        """
        Imports all the tables from worksheets within an Excel file.

        :param f_path: Path to the file
        :param wb_sheets: Worksheets to read.  Input as list i.e. ['Sales', 'Operations']
        :param table_coords: Excel references for top left and bottom right of table.  i.e. ['B7', 'Z20'].
        :param idx_cols: Number of index columns to use.  0 would be the first column and 3 would mean first 3 columns.
        :param dim_names: New dimensions names to add.  i.e. ['Department', 'Geography']
        :param dim_vals: Values for new dimensions.  i.e. ['Sales', 'North America']
        :param dim_names_coords: Excel references to get dimension names from the Excel.  i.e. ['A3', 'A4'].
        :param dim_coords: Excel references to get dimension values from the Excel.  i.e. ['B3', 'B4'].
        :return: self.data
        """
        dim_names = dim_names if isinstance(dim_names, list) else [dim_names]
        dim_vals = dim_vals if isinstance(dim_vals, list) else [dim_vals]
        dim_names_coords = dim_names_coords if isinstance(dim_names_coords, list) else [dim_names_coords]
        dim_coords = dim_coords if isinstance(dim_coords, list) else [dim_coords]
        wb = load_workbook(f_path, data_only=True)
        if wb_sheets == None:
            wb_sheets = wb.sheetnames
        if table_coords is None:
            for wsheet in wb_sheets:
                self.import_xl(f_path, wsheet)
                self.add_dimensions(["Work_Sheet"], [wsheet])
                self.add_block_to_data()
        else:
            for wsheet in wb_sheets:

                self.import_custom_xl(
                    f_path,
                    wsheet,
                    table_coords.copy(),
                    idx_cols,
                    dim_names.copy(),
                    dim_vals.copy(),
                    dim_names_coords.copy(),
                    dim_coords.copy(),
                    fill_index_na,
                )
                self.add_block_to_data()
                self.block = pd.DataFrame()

    def import_xl_directory(
        self,
        dir_path,
        xl_id=None,
        ws_name=0,
        table_coords=None,
        idx_cols=0,
        dim_names=None,
        dim_vals=None,
        dim_names_coords=None,
        dim_coords=None,
        fill_index_na=False,
    ):
        """
        Imports dimensions and table from a worksheet from all the Excel files (with or without identifiers) from a directory.

        :param dir_path: Path to the directory containing the files
        :param xl_id: String used to filter files to extract data, i.e. if file name is "Budget v3.xlsx" you could say 'v3.xlsx'
        :param ws_name: Worksheet to read.  Input as string i.e. 'Sales' or index.  Zero is default.
        :param table_coords: Excel references for top left and bottom right of table.  i.e. ['B7', 'Z20'].
        :param idx_cols: Number of index columns to use.  0 would be the first column and 3 would mean first 3 columns.
        :param dim_names: New dimensions names to add.  i.e. ['Department', 'Geography']
        :param dim_vals: Values for new dimensions.  i.e. ['Sales', 'North America']
        :param dim_names_coords: Excel references to get dimension names from the Excel.  i.e. ['A3', 'A4'].
        :param dim_coords: Excel references to get dimension values from the Excel.  i.e. ['B3', 'B4'].
        :return: self.data
        """
        dim_names = dim_names if isinstance(dim_names, list) else [dim_names]
        dim_vals = dim_vals if isinstance(dim_vals, list) else [dim_vals]
        dim_names_coords = dim_names_coords if isinstance(dim_names_coords, list) else [dim_names_coords]
        dim_coords = dim_coords if isinstance(dim_coords, list) else [dim_coords]
        file_list = pd.Series(os.listdir(dir_path))
        if xl_id != None:
            file_list = file_list[file_list.str.contains(xl_id, case=False, na=False)]
        for file in file_list:
            dir_path.replace("\\", "/")
            f_path = dir_path + "/" + file
            wb = load_workbook(f_path, data_only=True)
            # wb_sheets = wb.sheetnames
            wb_sheets = [ws_name]
            if table_coords is None:
                for wsheet in wb_sheets:
                    self.import_xl(f_path, wsheet)
                    self.add_dimensions(["Work_Sheet"], [wsheet])
                    self.add_block_to_data()
            else:
                for wsheet in wb_sheets:
                    self.import_custom_xl(
                        f_path,
                        ws_name,
                        table_coords.copy(),
                        idx_cols,
                        dim_names.copy(),
                        dim_vals.copy(),
                        dim_names_coords.copy(),
                        dim_coords.copy(),
                        fill_index_na,
                    )
                    self.add_block_to_data()
                    self.block = pd.DataFrame()

    def import_xl_directories(
        self,
        dir_path,
        xl_id=None,
        ws_name=0,
        table_coords=None,
        idx_cols=0,
        dim_names=None,
        dim_vals=None,
        dim_names_coords=None,
        dim_coords=None,
        fill_index_na=False,
    ):
        """
        Imports dimensions and table from a worksheet from all the Excel files (with or without identifiers) from a series of directories.

        :param dir_path: Path to the root directory containing the directories which contain the files.
        :param xl_id: String used to filter files to extract data, i.e. if file name is "Budget v3.xlsx" you could say 'v3.xlsx'
        :param ws_name: Worksheet to read.  Input as string i.e. 'Sales' or index.  Zero is default.
        :param table_coords: Excel references for top left and bottom right of table.  i.e. ['B7', 'Z20'].
        :param idx_cols: Number of index columns to use.  0 would be the first column and 3 would mean first 3 columns.
        :param dim_names: New dimensions names to add.  i.e. ['Department', 'Geography']
        :param dim_vals: Values for new dimensions.  i.e. ['Sales', 'North America']
        :param dim_names_coords: Excel references to get dimension names from the Excel.  i.e. ['A3', 'A4'].
        :param dim_coords: Excel references to get dimension values from the Excel.  i.e. ['B3', 'B4'].
        :return: self.data
        """
        dim_names = dim_names if isinstance(dim_names, list) else [dim_names]
        dim_vals = dim_vals if isinstance(dim_vals, list) else [dim_vals]
        dim_names_coords = dim_names_coords if isinstance(dim_names_coords, list) else [dim_names_coords]
        dim_coords = dim_coords if isinstance(dim_coords, list) else [dim_coords]
        dir_list = pd.Series(os.listdir(dir_path))
        dir_list = dir_list[~dir_list.str.contains("\.", na=False)]
        dir_path = dir_path.replace("\\", "/")
        for dir in dir_list:
            file_list = pd.Series(os.listdir(dir_path + "/" + dir))
            if xl_id != None:
                file_list = file_list[
                    (file_list.str.contains(xl_id, case=False, na=False)) &
                    (~file_list.str.contains('~', case=False, na=False))
                ]
            for file in file_list:
                f_path = dir_path + "/" + dir + "/" + file
                wb = load_workbook(f_path, data_only=True)
                # wb_sheets = wb.sheetnames
                wb_sheets = [ws_name]
                if table_coords is None:
                    for wsheet in wb_sheets:
                        self.import_xl(f_path, wsheet)
                        self.add_dimensions(["Work_Sheet"], [wsheet])
                        self.add_block_to_data()
                else:
                    for wsheet in wb_sheets:
                        self.import_custom_xl(
                            f_path,
                            ws_name,
                            table_coords.copy(),
                            idx_cols,
                            dim_names.copy(),
                            dim_vals.copy(),
                            dim_names_coords.copy(),
                            dim_coords.copy(),
                            fill_index_na,
                        )
                        self.add_block_to_data()
                        self.block = pd.DataFrame()

    def update_custom_xl(
        self,
        f_path=None,
        ws_name=0,
        table_coords=None,
        idx_cols=0,
        dim_names=None,
        dim_vals=None,
        dim_names_coords=None,
        dim_coords=None,
    ):
        """
        Custom update of existing data with mapping for data and dimensions from Excel Worksheet.  It will
        update a section of 'data', which has an identical dimension structure (excluding 'Data_Block' with new data.
        The new block's meta information will be referenced in the original blocks data to show history.

        :param f_path: Path to the Excel File.
        :param ws_name: Name or index number of the worksheet.  i.e. 'Accounts' or 2.
        :param table_coords: Excel references for top left and bottom right of table.  i.e. ['B7', 'Z20'].
        :param idx_cols: Number of index columns to use.  0 would be the first column and 3 would mean first 3 columns.
        :param dim_names: New dimensions names to add.  i.e. ['Department', 'Geography']
        :param dim_vals: Values for new dimensions.  i.e. ['Sales', 'North America']
        :param dim_names_coords: Excel references to get dimension names from the Excel.  i.e. ['A3', 'A4'].
        :param dim_coords: Excel references to get dimension values from the Excel.  i.e. ['B3', 'B4'].
        """
        table_coords = table_coords if isinstance(table_coords, list) else [table_coords]
        dim_names = dim_names if isinstance(dim_names, list) else [dim_names]
        dim_vals = dim_vals if isinstance(dim_vals, list) else [dim_vals]
        dim_names_coords = dim_names_coords if isinstance(dim_names_coords, list) else [dim_names_coords]
        dim_coords = dim_coords if isinstance(dim_coords, list) else [dim_coords]
        self.import_custom_xl(
            f_path,
            ws_name,
            table_coords.copy(),
            idx_cols,
            dim_names.copy(),
            dim_vals.copy(),
            dim_names_coords.copy(),
            dim_coords.copy(),
        )
        self._align_indicies()
        self.reorder_dimensions(self.data.index.names)
        data_nodb = self.data.droplevel("Data_Block")
        block_nodb = self.block.droplevel("Data_Block")
        count_iloc = range(0, len(block_nodb.index))
        for item, itemx in zip(block_nodb.index, count_iloc):
            idx_data = data_nodb.index.get_loc(item)
            for col in self.block.columns:
                if col not in self.data.columns:
                    self.data[col] = np.nan
                col_id = self.data.columns.get_loc(col)
                col_id_b = self.block.columns.get_loc(col)
                self.data.iloc[idx_data, col_id] = self.block.iloc[itemx, col_id_b]
        if isinstance(self.data.columns[0], dt.datetime):
            try:
                self.data.sort_index(1, inplace=True)
            except:
                pass
        # Add modified block number to modified_by in meta_block
        idx_db = list(self.data.index.names).index(
            "Data_Block"
        )  # get the int position of DB
        if isinstance(
            self.data.iloc[idx_data], pd.Series
        ):
            data_db_no = self.data.iloc[idx_data].name[idx_db]  # Series
        else:
            data_db_no = self.data.iloc[idx_data].index.values[0][idx_db]  # DF
        block_db_no = self.block.index.get_level_values("Data_Block")[0]
        mod_by = str(self.meta_block.loc[data_db_no, "modified_by"])
        mod_by = mod_by + "|" + str(block_db_no)
        self.meta_block.loc[data_db_no, "modified_by"] = mod_by
        # TODO: REFINE this.

    def update_xl_sheets(
        self,
        f_path,
        wb_sheets=None,
        table_coords=None,
        idx_cols=0,
        dim_names=None,
        dim_vals=None,
        dim_names_coords=None,
        dim_coords=None,
    ):
        """
        Updates all the tables from worksheets within an Excel file. It will
        update a section of 'data', which has an identical dimension structure (excluding 'Data_Block' with new data.
        The new block's meta information will be referenced in the original blocks data to show history.

        :param f_path: Path to the file
        :param wb_sheets: Worksheets to read.  Input as list i.e. ['Sales', 'Operations']
        :param table_coords: Excel references for top left and bottom right of table.  i.e. ['B7', 'Z20'].
        :param idx_cols: Number of index columns to use.  0 would be the first column and 3 would mean first 3 columns.
        :param dim_names: New dimensions names to add.  i.e. ['Department', 'Geography']
        :param dim_vals: Values for new dimensions.  i.e. ['Sales', 'North America']
        :param dim_names_coords: Excel references to get dimension names from the Excel.  i.e. ['A3', 'A4'].
        :param dim_coords: Excel references to get dimension values from the Excel.  i.e. ['B3', 'B4'].
        :return: self.data
        """
        dim_names = dim_names if isinstance(dim_names, list) else [dim_names]
        dim_vals = dim_vals if isinstance(dim_vals, list) else [dim_vals]
        dim_names_coords = dim_names_coords if isinstance(dim_names_coords, list) else [dim_names_coords]
        dim_coords = dim_coords if isinstance(dim_coords, list) else [dim_coords]
        wb = load_workbook(f_path, data_only=True)
        if wb_sheets == None:
            wb_sheets = wb.sheetnames
        if table_coords is None:
            for wsheet in wb_sheets:
                self.import_xl(f_path, wsheet)
                self.add_dimensions(["Work_Sheet"], [wsheet])
                self.add_block_to_data()
        else:
            for wsheet in wb_sheets:
                self.update_custom_xl(
                    f_path,
                    wsheet,
                    table_coords.copy(),
                    idx_cols,
                    dim_names.copy(),
                    dim_vals.copy(),
                    dim_names_coords.copy(),
                    dim_coords.copy(),
                )
                self.block = pd.DataFrame()

    def update_xl_directory(
        self,
        dir_path,
        xl_id=None,
        ws_name=0,
        table_coords=None,
        idx_cols=0,
        dim_names=None,
        dim_vals=None,
        dim_names_coords=None,
        dim_coords=None,
    ):
        """
        Updates dimensions and table from a worksheet from all the Excel files (with or without identifiers) from a directory. It will
        update a section of 'data', which has an identical dimension structure (excluding 'Data_Block' with new data.
        The new block's meta information will be referenced in the original blocks data to show history.

        :param dir_path: Path to the directory containing the files
        :param xl_id: String used to filter files to extract data, i.e. if file name is "Budget v3.xlsx" you could say 'v3.xlsx'
        :param ws_name: Worksheet to read.  Input as string i.e. 'Sales' or index.  Zero is default.
        :param table_coords: Excel references for top left and bottom right of table.  i.e. ['B7', 'Z20'].
        :param idx_cols: Number of index columns to use.  0 would be the first column and 3 would mean first 3 columns.
        :param dim_names: New dimensions names to add.  i.e. ['Department', 'Geography']
        :param dim_vals: Values for new dimensions.  i.e. ['Sales', 'North America']
        :param dim_names_coords: Excel references to get dimension names from the Excel.  i.e. ['A3', 'A4'].
        :param dim_coords: Excel references to get dimension values from the Excel.  i.e. ['B3', 'B4'].
        :return: self.data
        """
        dim_names = dim_names if isinstance(dim_names, list) else [dim_names]
        dim_vals = dim_vals if isinstance(dim_vals, list) else [dim_vals]
        dim_names_coords = dim_names_coords if isinstance(dim_names_coords, list) else [dim_names_coords]
        dim_coords = dim_coords if isinstance(dim_coords, list) else [dim_coords]
        file_list = pd.Series(os.listdir(dir_path))
        if xl_id != None:
            file_list = file_list[file_list.str.contains(xl_id, case=False, na=False)]
        for file in file_list:
            dir_path.replace("\\", "/")
            f_path = dir_path + "/" + file
            wb_sheets = [ws_name]
            if table_coords is None:
                for wsheet in wb_sheets:
                    self.import_xl(f_path, wsheet)
                    self.add_dimensions(["Work_Sheet"], [wsheet])
                    self.add_block_to_data()
            else:
                for wsheet in wb_sheets:
                    self.update_custom_xl(
                        f_path,
                        ws_name,
                        table_coords.copy(),
                        idx_cols,
                        dim_names.copy(),
                        dim_vals.copy(),
                        dim_names_coords.copy(),
                        dim_coords.copy(),
                    )
                    self.block = pd.DataFrame()

    def update_xl_directories(
        self,
        dir_path,
        xl_id=None,
        ws_name=0,
        table_coords=None,
        idx_cols=0,
        dim_names=None,
        dim_vals=None,
        dim_names_coords=None,
        dim_coords=None,
    ):
        """
        Imports dimensions and table from a worksheet from all the Excel files (with or without identifiers) from a series of directories. It will
        update a section of 'data', which has an identical dimension structure (excluding 'Data_Block' with new data.
        The new block's meta information will be referenced in the original blocks data to show history.

        :param dir_path: Path to the root directory containing the directories which contain the files.
        :param xl_id: String used to filter files to extract data, i.e. if file name is "Budget v3.xlsx" you could say 'v3.xlsx'
        :param ws_name: Worksheet to read.  Input as string i.e. 'Sales' or index.  Zero is default.
        :param table_coords: Excel references for top left and bottom right of table.  i.e. ['B7', 'Z20'].
        :param idx_cols: Number of index columns to use.  0 would be the first column and 3 would mean first 3 columns.
        :param dim_names: New dimensions names to add.  i.e. ['Department', 'Geography']
        :param dim_vals: Values for new dimensions.  i.e. ['Sales', 'North America']
        :param dim_names_coords: Excel references to get dimension names from the Excel.  i.e. ['A3', 'A4'].
        :param dim_coords: Excel references to get dimension values from the Excel.  i.e. ['B3', 'B4'].
        :return: self.data
        """
        dim_names = dim_names if isinstance(dim_names, list) else [dim_names]
        dim_vals = dim_vals if isinstance(dim_vals, list) else [dim_vals]
        dim_names_coords = dim_names_coords if isinstance(dim_names_coords, list) else [dim_names_coords]
        dim_coords = dim_coords if isinstance(dim_coords, list) else [dim_coords]
        dir_list = pd.Series(os.listdir(dir_path))
        dir_list = dir_list[~dir_list.str.contains("\.", na=False)]
        dir_path = dir_path.replace("\\", "/")
        for dir in dir_list:
            file_list = pd.Series(os.listdir(dir_path + "/" + dir))
            if xl_id != None:
                file_list = file_list[
                    file_list.str.contains(xl_id, case=False, na=False)
                ]
            for file in file_list:
                f_path = dir_path + "/" + dir + "/" + file
                wb_sheets = [ws_name]
                if table_coords is None:
                    for wsheet in wb_sheets:
                        self.import_xl(f_path, wsheet)
                        self.add_dimensions(["Work_Sheet"], [wsheet])
                        self.add_block_to_data()
                else:
                    for wsheet in wb_sheets:
                        self.update_custom_xl(
                            f_path,
                            ws_name,
                            table_coords.copy(),
                            idx_cols,
                            dim_names.copy(),
                            dim_vals.copy(),
                            dim_names_coords.copy(),
                            dim_coords.copy(),
                        )
                        self.block = pd.DataFrame()

    def add_block_to_data(self):
        """Takes the block data object, arranges the index and adds it to the data object.  Even if the indexes don't
        match, this will fill in the missing pieces."""
        if self.data.index.names != [None]:
            self._align_indicies()
            self.reorder_dimensions(self.data.index.names)
        self.data = pd.concat([self.data, self.block])

    help_add_dimensions = "Add dimensions and values with lists ['???',...]"

    def add_dimensions(
        self, new_dimensions, dim_values_to_add, col_num=1, data_obj="data"
    ):
        """
        Append a dimension to the index with a new name, values and where to place it.

        :param new_dimensions: Name of the new dimension or dimensions.  String value or list i.e. 'Department' or
         ['Department', 'Region']
        :param dim_values_to_add: Values of the new dimension or dimensions.  String value or list i.e. 'Sales' or
         ['Sales', 'EMEA']
        :param col_num: Where in the index to place the new dimension.  1 indicates 1 column from the left.
        :param data_obj: Which data object you want o effect.  Available - 'block', 'slice', 'consolidation', 'function_result', 'data'
        :return: data_obj
        """
        new_dimensions = new_dimensions if isinstance(new_dimensions, list) else [new_dimensions]
        dim_values_to_add = dim_values_to_add if isinstance(dim_values_to_add, list) else [dim_values_to_add]
        if data_obj == "block":
            dfi = self.block.index.to_frame()
            for dim_name, dim_value in zip(new_dimensions, dim_values_to_add):
                dfi.insert(col_num, dim_name, dim_value)
            self.block.index = pd.MultiIndex.from_frame(dfi)
        elif data_obj == "slice":
            dfi = self.slice.index.to_frame()
            for dim_name, dim_value in zip(new_dimensions, dim_values_to_add):
                dfi.insert(col_num, dim_name, dim_value)
            self.slice.index = pd.MultiIndex.from_frame(dfi)
        elif data_obj == "consolidation":
            dfi = self.consolidation.index.to_frame()
            for dim_name, dim_value in zip(new_dimensions, dim_values_to_add):
                dfi.insert(col_num, dim_name, dim_value)
            self.consolidation.index = pd.MultiIndex.from_frame(dfi)
        elif data_obj == "function_result":
            dfi = self.function_result.index.to_frame()
            for dim_name, dim_value in zip(new_dimensions, dim_values_to_add):
                dfi.insert(col_num, dim_name, dim_value)
            self.function_result.index = pd.MultiIndex.from_frame(dfi)
        else:
            if (
                len(new_dimensions) == len(dim_values_to_add)
                and isinstance(dim_values_to_add[0], list) is False
            ):
                dfi = self.data.index.to_frame()
                for dim_name, dim_value in zip(new_dimensions, dim_values_to_add):
                    dfi.insert(col_num, dim_name, dim_value)
                self.data.index = pd.MultiIndex.from_frame(dfi)
            else:
                dfi = self.data.index.to_frame()
                index_counts = self.data._count_level(0).iloc[:, 0]
                index_counts = index_counts.reindex(
                    dfi.index.get_level_values("Data_Block").unique()
                )
                for new_dim_name in new_dimensions:
                    dim_value_index = new_dimensions.index(new_dim_name)
                    dim_value_list = []
                    counter = 0
                    dim_value_loop = dim_values_to_add[dim_value_index]
                    for idx_value in index_counts.index:
                        dim_value = (
                            dim_value_loop[counter]
                            if len(dim_value_loop) > 1
                            else dim_value_loop[0]
                        )
                        number_of_iter = list(range(0, index_counts.loc[idx_value]))
                        counter += 1
                        for item in number_of_iter:
                            dim_value_list.append(dim_value)
                    dim_value_list = pd.Series(dim_value_list)
                    dim_value_list.name = new_dim_name
                    dfi.insert(1, new_dim_name, dim_value_list.values)
            self.data.index = pd.MultiIndex.from_frame(dfi)

            # TODO: FIX add dimension to data for multiple. work on ordering when adding

    help_merge_dim_from_accts = (
        "One-to-One add of a dimension from an accounts based on existing dimension"
    )

    def merge_dim_from_accts(self, dim_set, base_dim, new_dims, data_obj="data"):
        """
        Add dimension based on account data object. If you have a chart of accounts you could add the account number
        to the Line Item if you have it.

        :param dim_set: Section of 'accounts' data object to take new dimensions.
        :param base_dim: Existing dimension in the index on which to merge the new dimension
        :param new_dims: Column from the accounts data object table to add to the dataframe
        :param data_obj: Which data object you want o effect.  Available - 'block', 'data', 'slice', 'consolidation',
         'function_result'
        :return: data_object
        """
        dim_df = self.accounts.loc[dim_set].copy()
        dim_df.dropna(1, how="all", inplace=True)
        new_dims = [new_dims] if isinstance(new_dims, str) else new_dims
        if data_obj == "block":
            dfi = self.block.index.to_frame()
            dfi.reset_index(drop=True, inplace=True)
            new_dims.append(base_dim)
            dim_df = dim_df.loc[:, new_dims]
            dim_df.set_index(base_dim, inplace=True)
            dfi = dfi.join(dim_df, base_dim, how="inner")
            self.block.index = pd.MultiIndex.from_frame(dfi)
        elif data_obj == "slice":
            dfi = self.slice.index.to_frame()
            dfi.reset_index(drop=True, inplace=True)
            new_dims.append(base_dim)
            dim_df = dim_df.loc[:, new_dims]
            dim_df.set_index(base_dim, inplace=True)
            dfi = dfi.join(dim_df, base_dim, how="inner")
            self.slice.index = pd.MultiIndex.from_frame(dfi)
        elif data_obj == "consolidation":
            dfi = self.consolidation.index.to_frame()
            dfi.reset_index(drop=True, inplace=True)
            new_dims.append(base_dim)
            dim_df = dim_df.loc[:, new_dims]
            dim_df.set_index(base_dim, inplace=True)
            dfi = dfi.join(dim_df, base_dim, how="inner")
            self.consolidation.index = pd.MultiIndex.from_frame(dfi)
        elif data_obj == "variance":
            dfi = self.variance.index.to_frame()
            dfi.reset_index(drop=True, inplace=True)
            new_dims.append(base_dim)
            dim_df = dim_df.loc[:, new_dims]
            dim_df.set_index(base_dim, inplace=True)
            dfi = dfi.join(dim_df, base_dim, how="inner")
            self.variance.index = pd.MultiIndex.from_frame(dfi)
        elif data_obj == "function_result":
            dfi = self.function_result.index.to_frame()
            dfi.reset_index(drop=True, inplace=True)
            new_dims.append(base_dim)
            dim_df = dim_df.loc[:, new_dims]
            dim_df.set_index(base_dim, inplace=True)
            dfi = dfi.join(dim_df, base_dim, how="inner")
            self.function_result.index = pd.MultiIndex.from_frame(dfi)
        else:
            dfi = self.data.index.to_frame()
            dfi.reset_index(drop=True, inplace=True)
            new_dims.append(base_dim)
            dim_df = dim_df.loc[:, new_dims]
            dim_df.set_index(base_dim, inplace=True)
            dfi = dfi.join(dim_df, base_dim, how="inner")
            self.data.index = pd.MultiIndex.from_frame(dfi)

    help_merge_dim_from_xl = (
        "One-to-One add of a dimension from an excel list based on existing dimension"
    )

    def merge_dim_from_xl(self, fpath, ws_name, base_dim, new_dims, data_obj="data"):
        """
        Add dimension based on an table from an Excel file. If you have a chart of accounts you could add the account number
        to the Line Item if you have it.

        :param base_dim: Existing dimension in the index on which to merge the new dimension
        :param new_dims: Column from the accounts data object table to add to the dataframe
        :param data_obj: Which data object you want o effect.  Available - 'block', 'data', 'slice', 'consolidation',
         'function_result'
        :return: data_object
        """
        dim_df = pd.read_excel(fpath, sheet_name=ws_name, index_col=None)
        new_dims = [new_dims] if isinstance(new_dims, str) else new_dims
        if data_obj == "block":
            dfi = self.block.index.to_frame()
            dfi.reset_index(drop=True, inplace=True)
            new_dims.append(base_dim)
            dim_df = dim_df.loc[:, new_dims]
            dim_df.set_index(base_dim, inplace=True)
            dfi = dfi.join(dim_df, base_dim, how="inner")
            self.block.index = pd.MultiIndex.from_frame(dfi)
        elif data_obj == "slice":
            dfi = self.slice.index.to_frame()
            dfi.reset_index(drop=True, inplace=True)
            new_dims.append(base_dim)
            dim_df = dim_df.loc[:, new_dims]
            dim_df.set_index(base_dim, inplace=True)
            dfi = dfi.join(dim_df, base_dim, how="inner")
            self.slice.index = pd.MultiIndex.from_frame(dfi)
        elif data_obj == "consolidation":
            dfi = self.consolidation.index.to_frame()
            dfi.reset_index(drop=True, inplace=True)
            new_dims.append(base_dim)
            dim_df = dim_df.loc[:, new_dims]
            dim_df.set_index(base_dim, inplace=True)
            dfi = dfi.join(dim_df, base_dim, how="inner")
            self.consolidation.index = pd.MultiIndex.from_frame(dfi)
        elif data_obj == "variance":
            dfi = self.variance.index.to_frame()
            dfi.reset_index(drop=True, inplace=True)
            new_dims.append(base_dim)
            dim_df = dim_df.loc[:, new_dims]
            dim_df.set_index(base_dim, inplace=True)
            dfi = dfi.join(dim_df, base_dim, how="inner")
            self.variance.index = pd.MultiIndex.from_frame(dfi)
        elif data_obj == "function_result":
            dfi = self.function_result.index.to_frame()
            dfi.reset_index(drop=True, inplace=True)
            new_dims.append(base_dim)
            dim_df = dim_df.loc[:, new_dims]
            dim_df.set_index(base_dim, inplace=True)
            dfi = dfi.join(dim_df, base_dim, how="inner")
            self.function_result.index = pd.MultiIndex.from_frame(dfi)
        else:
            dfi = self.data.index.to_frame()
            dfi.reset_index(drop=True, inplace=True)
            new_dims.append(base_dim)
            dim_df = dim_df.loc[:, new_dims]
            dim_df.set_index(base_dim, inplace=True)
            dfi = dfi.join(dim_df, base_dim, how="inner")
            self.data.index = pd.MultiIndex.from_frame(dfi)

    def reorder_dimensions(self, new_order, data_obj="data"):
        """
        Change the order of the dimensions of the index.

        :param new_order: List input i.e. ['Department', 'Region', 'Data_Block', 'Line Item']
        :param data_obj: Which data object you want o effect.  Available - 'block', 'data', 'slice', 'consolidation',
         'function_result', 'variance'
        :return: data_obj
        """
        if data_obj == "block":
            self.block = self.block.reorder_levels(new_order)
        elif data_obj == "data":
            self.data = self.data.reorder_levels(new_order)
        elif data_obj == "function_result":
            self.function_result= self.function_result.reorder_levels(new_order)
        elif data_obj == "variance":
            self.variance = self.variance.reorder_levels(new_order)
        elif data_obj == "consolidation":
            self.consolidation = self.consolidation.reorder_levels(new_order)
        else:
            self.slice = self.slice.reorder_levels(new_order)

    def combine_dimensions(self, combine_order, data_obj="data"):
        if data_obj == "block":
            dfi = self.block.index.to_frame()
            dfi.reset_index(drop=True, inplace=True)
            s0 = dfi[combine_order[0]]
            for item in combine_order[1:]:
                s0 = s0.combine_first(dfi.loc[:, item])
            dfi[combine_order[0]] = s0
            self.block.index = pd.MultiIndex.from_frame(dfi)
        elif data_obj == "data":
            dfi = self.data.index.to_frame()
            dfi.reset_index(drop=True, inplace=True)
            s0 = dfi[combine_order[0]]
            for item in combine_order[1:]:
                s0 = s0.combine_first(dfi.loc[:, item])
            dfi[combine_order[0]] = s0
            self.data.index = pd.MultiIndex.from_frame(dfi)

    def rename_dimensions(self, dim_list, data_obj="data"):
        """
        Give new names to one or all of the dimension names.  I.e. if dimensions are ['Department', 'Region', 'Data_Block']
        it can be changed to ['Department', 'Geography', 'Data_Block']
        :param dim_list: List of all the dimension names with the new names included
        :param data_obj: Which data object you want o effect.  Available - 'block', 'data', 'slice', 'consolidation',
         'function_result', 'variance'
        :return: data object
        """
        if data_obj == "data":
            self.data.index.names = dim_list
        elif data_obj == "block":
            self.block.index.names = dim_list
        elif data_obj == "slice":
            self.slice.index.names = dim_list
        elif data_obj == "variance":
            self.variance.index.names = dim_list
        elif data_obj == "consolidation":
            self.consolidation.index.names = dim_list
        else:
            self.function_result.index.names = dim_list

    def rename_dim_item(self, dim, old, new, data_obj="data"):
        """
        Give a new name to an existing name in a dimension.  I.e. if 'USA" is to be changed to 'North America'

        :param dim: String value of which dimension contains the item you want to change.  I.e. 'Region'
        :param old: String value of the item to be changed.  I.e. 'USA'
        :param new: String value of the new name of the item.  I.e. 'North America'
        :param data_obj: Which data object you want o effect.  Available - 'block', 'data', 'slice', 'consolidation',
         'function_result', 'variance'
        :return: data object
        """
        if data_obj == "data":
            idx = self.data.index.to_frame()
            idx.fillna("nval", inplace=True)
            idx[dim] = idx[dim].str.replace(old, new)
            self.data.index = pd.MultiIndex.from_frame(idx)
        elif data_obj == "block":
            idx = self.block.index.to_frame()
            idx.fillna("nval", inplace=True)
            idx[dim] = idx[dim].str.replace(old, new)
            self.block.index = pd.MultiIndex.from_frame(idx)
        elif data_obj == "accounts":
            idx = self.accounts.index.to_frame()
            idx.fillna("nval", inplace=True)
            idx['dim_set'] = idx['dim_set'].str.replace(old, new)
            self.accounts.index = pd.MultiIndex.from_frame(idx)
        elif data_obj == "slice":
            idx = self.slice.index.to_frame()
            idx.fillna("nval", inplace=True)
            idx[dim] = idx[dim].str.replace(old, new)
            self.slice.index = pd.MultiIndex.from_frame(idx)
        elif data_obj == "function_result":
            idx = self.function_result.index.to_frame()
            idx.fillna("nval", inplace=True)
            idx[dim] = idx[dim].str.replace(old, new)
            self.function_result.index = pd.MultiIndex.from_frame(idx)
        elif data_obj == "consolidation":
            idx = self.consolidation.index.to_frame()
            idx.fillna("nval", inplace=True)
            idx[dim] = idx[dim].str.replace(old, new)
            self.consolidation.index = pd.MultiIndex.from_frame(idx)
        else:
            idx = self.variance.index.to_frame()
            idx.fillna("nval", inplace=True)
            idx[dim] = idx[dim].str.replace(old, new)
            self.variance.index = pd.MultiIndex.from_frame(idx)

    def dim_to_date(self, dim, data_obj="data"):
        """
        Attempts to change a dimension or column from object to datetime.

        :param dim: String value of which dimension contains the item you want to change.  I.e. 'Region'
        :param data_obj: Which data object you want o effect.  Available - 'block', 'data', 'slice', 'consolidation',
         'function_result', 'variance'
        """
        if data_obj == "data":
            idx = self.data.index.to_frame()
            idx.fillna("nval", inplace=True)
            idx[dim] = pd.to_datetime(idx[dim])
            self.data.index = pd.MultiIndex.from_frame(idx)
        elif data_obj == "block":
            idx = self.block.index.to_frame()
            idx.fillna("nval", inplace=True)
            idx[dim] = pd.to_datetime(idx[dim])
            self.block.index = pd.MultiIndex.from_frame(idx)
        elif data_obj == "accounts":
            idx = self.accounts.index.to_frame()
            idx.fillna("nval", inplace=True)
            idx[dim] = pd.to_datetime(idx[dim])
            self.accounts.index = pd.MultiIndex.from_frame(idx)
        elif data_obj == "slice":
            idx = self.slice.index.to_frame()
            idx.fillna("nval", inplace=True)
            idx[dim] = pd.to_datetime(idx[dim])
            self.slice.index = pd.MultiIndex.from_frame(idx)
        elif data_obj == "function_result":
            idx = self.function_result.index.to_frame()
            idx.fillna("nval", inplace=True)
            idx[dim] = pd.to_datetime(idx[dim])
            self.function_result.index = pd.MultiIndex.from_frame(idx)
        elif data_obj == "consolidation":
            idx = self.consolidation.index.to_frame()
            idx.fillna("nval", inplace=True)
            idx[dim] = pd.to_datetime(idx[dim])
            self.consolidation.index = pd.MultiIndex.from_frame(idx)
        else:
            idx = self.variance.index.to_frame()
            idx.fillna("nval", inplace=True)
            idx[dim] = pd.to_datetime(idx[dim])
            self.variance.index = pd.MultiIndex.from_frame(idx)

    def reorder_index_dim(
        self, new_order, dim=None, axis_target="index", data_obj="data"
    ):
        """
        Change the order of the item within an index.  I.e. if existing the Line Item dimension has an existing order of
        ['Office Supplies', 'Rent', 'Payroll'], change it to ['Payroll', 'Rent', 'Office Supplies'].

        :param new_order: List values of the new order.  Should contain all the values.
        I.e. ['Payroll', 'Rent', 'Office Supplies']
        :param dim: Dimension to be reordered
        :param axis_target: Index or the columns.  I.e. 'index' or 'columns'
        :param data_obj: Which data object you want o effect.  Available - 'block', 'data', 'slice', 'consolidation',
         'function_result', 'variance'
        :return:
        """
        if data_obj == "data":
            index_items = list(set(self.data.index.get_level_values(dim)))
            for item_no in new_order:
                index_items.remove(item_no)
            new_order.reverse()
            for item_no in new_order:
                index_items.insert(0, item_no)
            self.data = self.data.reindex(
                index_items, axis=axis_target, level=dim, fill_value="nval"
            )
        elif data_obj == "block":
            index_items = list(set(self.block.index.get_level_values(dim)))
            for item_no in new_order:
                index_items.remove(item_no)
            new_order.reverse()
            for item_no in new_order:
                index_items.insert(0, item_no)
            self.block = self.block.reindex(
                index_items, axis=axis_target, level=dim, fill_value="nval"
            )
        elif data_obj == "slice":
            index_items = list(set(self.slice.index.get_level_values(dim)))
            for item_no in new_order:
                index_items.remove(item_no)
            new_order.reverse()
            for item_no in new_order:
                index_items.insert(0, item_no)
            self.slice = self.slice.reindex(
                index_items, axis=axis_target, level=dim, fill_value="nval"
            )
        elif data_obj == "function_result":
            index_items = list(set(self.function_result.index.get_level_values(dim)))
            for item_no in new_order:
                index_items.remove(item_no)
            new_order.reverse()
            for item_no in new_order:
                index_items.insert(0, item_no)
            self.function_result= self.function_result.reindex(
                index_items, axis=axis_target, level=dim, fill_value="nval"
            )
        elif data_obj == "consolidation":
            index_items = list(set(self.consolidation.index.get_level_values(dim)))
            for item_no in new_order:
                index_items.remove(item_no)
            new_order.reverse()
            for item_no in new_order:
                index_items.insert(0, item_no)
            self.consolidation = self.consolidation.reindex(
                index_items, axis=axis_target, level=dim, fill_value="nval"
            )
        else:
            index_items = list(set(self.variance.index.get_level_values(dim)))
            for item_no in new_order:
                index_items.remove(item_no)
            new_order.reverse()
            for item_no in new_order:
                index_items.insert(0, item_no)
            self.variance = self.variance.reindex(
                index_items, axis=axis_target, level=dim, fill_value="nval"
            )

    def move_dims_to_col(self, dims, data_obj="data"):
        """
        Move a dimension from the index to a column in the data.

        :param dims: Dimension to drop.  I.e. 'Department'
        :param data_obj: Which data object you want o effect.  Available - 'block', 'data', 'slice', 'consolidation',
         'function_result', 'variance'
        :return:
        """
        if data_obj == "data":
            self.data.reset_index(dims, inplace=True)
        elif data_obj == "block":
            self.block.reset_index(dims, inplace=True)
        elif data_obj == "slice":
            self.slice.reset_index(dims, inplace=True)
        elif data_obj == "variance":
            self.variance.index.reset_index(dims, inplace=True)
        elif data_obj == "consolidation":
            self.consolidation.reset_index(
                dims, inplace=True
            )
        else:
            self.function_result.reset_index(dims, inplace=True)

    def move_col_to_dims(self, dims, data_obj="data"):
        """
        Move a dimension from the columns in the data to the index.

        :param dims: Dimension to drop.  I.e. 'Department'
        :param data_obj: Which data object you want o effect.  Available - 'block', 'data', 'slice', 'consolidation',
         'function_result', 'variance'
        :return:
        """
        if data_obj == "data":
            self.data.set_index(dims, append=True, inplace=True)
        elif data_obj == "block":
            self.block.set_index(dims, append=True, inplace=True)
        elif data_obj == "slice":
            self.slice.set_index(dims, append=True, inplace=True)
        elif data_obj == "variance":
            self.variance.set_index(dims, append=True, inplace=True)
        elif data_obj == "consolidation":
            self.consolidation.set_index(dims, append=True, inplace=True)
        else:
            self.function_result.set_index(dims, append=True, inplace=True)

    def make_records(self, data_obj="data"):
        """
        Change the data object to a records format rather than a table.  Works bes on single level columns.

        :param data_obj: Which data object you want o effect. Default: 'data'.  Available - 'block', 'data', 'slice',
         'consolidation', 'function_result', 'variance'
        """
        if data_obj == "data":
            self.function_result = self.data.copy().stack()
        elif data_obj == "slice":
            self.function_result = self.slice.copy().stack()
        elif data_obj == "block":
            self.function_result = self.block.copy().stack()
        elif data_obj == "variance":
            self.function_result = self.variance.copy().stack()
        elif data_obj == "consolidation":
            self.function_result = self.consolidation.copy().stack()
        elif data_obj == "function_result":
            self.block= self.function_result.copy().stack()
        else:
            print("no object")

    def drop_dimension(self, dimension_drop, data_obj="data"):
        """
        Remove a dimension from the index.

        :param dimension_drop: Dimension to drop.  I.e. 'Department'
        :param data_obj: Which data object you want o effect.  Available - 'block', 'data', 'slice', 'consolidation',
         'function_result', 'variance'
        :return:
        """
        if data_obj == "data":
            self.data.index = self.data.index.droplevel(dimension_drop)
        elif data_obj == "block":
            self.block.index = self.block.index.droplevel(dimension_drop)
        elif data_obj == "slice":
            self.slice.index = self.slice.index.droplevel(dimension_drop)
        elif data_obj == "variance":
            self.variance.index = self.variance.index.droplevel(dimension_drop)
        elif data_obj == "consolidation":
            self.consolidation.index = self.consolidation.index.droplevel(
                dimension_drop
            )
        else:
            self.function_result.index = self.function_result.index.droplevel(dimension_drop)

    def get_block_info(self, db_id):
        """
        Retrieve the meta information from a data block.  Just plug in the data block number.

        :param db_id: Interger of the data block. I.e. x.get_block_info(49593949)
        :return: Meta information for Data_Block.
        """
        meta_info = self.meta_block.loc[db_id]
        return meta_info

    def consol_dimension(self, dims="Data_Block", data_obj="data"):
        """
        Consolidates a data object on a certain dimension of the data object.

        :param dim: Dimension to consolidate on.  I.e. 'Department' or 'SalesPerson'.
        :param data_obj: Which data object you want o effect. Default: 'data'.  Available - 'block', 'data', 'slice', 'consolidation',
         'function_result', 'variance'
        :return: self.consolidation
        """
        last_dim = self.data.index.names[-1]
        if data_obj == "data":
            self.consolidation = self.data.groupby(level=dims, sort=False).sum()
        elif data_obj == "slice":
            self.consolidation = self.slice.groupby(level=dims, sort=False).sum()
        elif data_obj == "function_result":
            self.consolidation = self.function_result.groupby(level=dims, sort=False).sum()
        elif data_obj == "block":
            self.consolidation = self.block.groupby(level=dims, sort=False).sum()
        elif data_obj == "variance":
            self.consolidation = self.variance.groupby(level=dims, sort=False).sum()
        else:
            self.consolidation = self.consolidation.groupby(level=dims, sort=False).sum()
        return self.consolidation

    def slice_data(
        self,
        dims=None,
        dim_values=None,
        col_range=None,
        col_list=None,
        data_obj="data",
        append_to=False,
    ):
        """
        Slice and dice data based on dimensions for the index and ranges or lists for the columns.

        :param dims: List object (even for one item) of dimensions.  I.e. ['Department', 'Line Item']
        :param dim_values: List and nested list object (even for one item) of items on which to slice.
            I.e. ['Operations', ['Network Costs', 'Payroll']]
        :param col_range: List object of start and end point of range. I.e. ['2022-06-30', '2022-12-31'] for datetime
        :param col_list: List object columns. I.e. ['2022-06-30', '2022-09-30', '2022-12-31'] for datetime
        :param data_obj: Which data object you want o effect.  Available - 'data', 'slice'
        :param append_to: Add data to existing slice data object.
        :return: self.slice
        """
        dims = dims if isinstance(dims, list) else [dims]
        dim_values = dim_values if isinstance(dim_values, list) else [dim_values]

        if data_obj == "data":
            col_search_dims = []
            col_search_values = []
            col_search_dims = [
                dim
                for dim in dims
                if dim not in list(self.data.index.names)
                and dim in list(self.data.columns)
            ]
            dims_idx = [dims.index(col) for col in col_search_dims]
            col_search_values = [dim_values[idx] for idx in dims_idx]
            dims = [dim for dim in dims if dim not in col_search_dims]
            dim_values = [dim for dim in dim_values if dim not in col_search_values]
            if append_to == False:
                self.slice = pd.DataFrame()
            if dims == None or dims == [None] or dims == []:
                self.slice = self.data.copy()
            else:
                slice_matrix = pd.DataFrame(columns=dims)
                if dim_values == None or dim_values == [None]:
                    dim_values = []
                    if isinstance(self.data.index, pd.MultiIndex):
                        for item in self.data.index.levels:
                            dim_values.append(list(item))
                    else:
                        dim_values = self.data.index.values
                count_row = 0
                for item, dim_name in zip(dim_values, dims):
                    sub_count_row = 0
                    if isinstance(item, list):
                        for subitem in item:
                            slice_matrix.loc[sub_count_row, dim_name] = subitem
                            sub_count_row += 1
                    else:
                        slice_matrix.loc[count_row, dim_name] = item

                iter_matrix = pd.DataFrame()
                df_count = pd.DataFrame(
                    index=["dim_count"], columns=slice_matrix.columns
                )
                for col in slice_matrix.columns:
                    dm = len(slice_matrix[col].dropna())
                    df_count.loc["dim_count", col] = dm
                for target_col in df_count.columns:
                    for target_row in slice_matrix[target_col].dropna().index:
                        target_column = slice_matrix[target_col].dropna()
                        target_dim_item = target_column.loc[target_row]
                        df_rest_list = list(slice_matrix.columns)
                        df_rest_list.remove(target_col)
                        df_rest = slice_matrix.loc[:, df_rest_list]
                        x = 1
                        for item in df_rest.columns:
                            x = len(slice_matrix[item].dropna()) * x
                        df_item = pd.DataFrame(
                            index=range(0, x), columns=slice_matrix.columns
                        )
                        df_item[target_col] = target_dim_item
                        for rest_col in df_rest.columns:
                            rest_column = df_rest[rest_col].dropna()
                            rest_item_count = df_count[rest_col][0]
                            rest_multiplier = (
                                int(len(df_item.index) / rest_item_count) - 1
                            )
                            rest_column_insert = rest_column.copy()
                            for mlper in range(0, rest_multiplier):
                                rest_column_insert = rest_column_insert.append(
                                    rest_column
                                )
                            rest_column_insert.index = df_item.index
                            df_item[rest_col] = rest_column_insert

                        iter_matrix = pd.concat(
                            [iter_matrix, df_item], ignore_index=True
                        )

                iter_matrix = (
                    iter_matrix.drop_duplicates()
                    .dropna(how="any")
                    .reset_index(drop=True)
                )
                for row in iter_matrix.index:
                    slice_terms = tuple(iter_matrix.loc[row].values)
                    if isinstance(self.data.index, pd.MultiIndex):
                        single_slice = self.data.xs(
                            slice_terms, level=dims, drop_level=False
                        )
                    else:
                        for dim in dims:
                            for item_s in dim_values:
                                single_slice = self.data.loc[item_s]
                    self.slice = pd.concat([self.slice, single_slice])
                    if isinstance(self.data.index, pd.Index):
                        self.slice.drop_duplicates(inplace=True)
                self.slice.sort_index(inplace=True)
            # Column Search
            if col_search_dims != [] and col_search_values != []:
                col_slice = pd.DataFrame()
                col_slice_df = self.slice.copy()
                for dim in col_search_dims:
                    for val in col_search_values:
                        if isinstance(val, list):
                            for sub_val in val:
                                single_slice = col_slice_df[
                                    col_slice_df[dim] == sub_val
                                ]
                                col_slice = pd.concat([col_slice, single_slice])
                        else:
                            single_slice = col_slice_df[col_slice_df[dim] == val]
                            col_slice = pd.concat([col_slice, single_slice])
                    col_slice_df = col_slice
                self.slice = col_slice
            # Column axis slicing
            if col_range != None:
                self.slice = self.slice.loc[:, col_range[0] : col_range[1]]
            if col_list != None:
                self.slice = self.slice.loc[:, col_list]
            return self.slice

        if data_obj == "consolidation":
            col_search_dims = []
            col_search_values = []
            col_search_dims = [
                dim
                for dim in dims
                if dim not in list(self.consolidation.index.names)
                and dim in list(self.consolidation.columns)
            ]
            dims_idx = [dims.index(col) for col in col_search_dims]
            col_search_values = [dim_values[idx] for idx in dims_idx]
            dims = [dim for dim in dims if dim not in col_search_dims]
            dim_values = [dim for dim in dim_values if dim not in col_search_values]
            if append_to == False:
                self.slice = pd.DataFrame()
            if dims == None or dims == [None] or dims == []:
                self.slice = self.consolidation.copy()
            else:
                slice_matrix = pd.DataFrame(columns=dims)
                if dim_values == None or dim_values == [None]:
                    dim_values = []
                    if isinstance(self.consolidation.index, pd.MultiIndex):
                        for item in self.consolidation.index.levels:
                            dim_values.append(list(item))
                    else:
                        dim_values = self.consolidation.index.values
                count_row = 0
                for item, dim_name in zip(dim_values, dims):
                    sub_count_row = 0
                    if isinstance(item, list):
                        for subitem in item:
                            slice_matrix.loc[sub_count_row, dim_name] = subitem
                            sub_count_row += 1
                    else:
                        slice_matrix.loc[count_row, dim_name] = item

                iter_matrix = pd.DataFrame()
                df_count = pd.DataFrame(
                    index=["dim_count"], columns=slice_matrix.columns
                )
                for col in slice_matrix.columns:
                    dm = len(slice_matrix[col].dropna())
                    df_count.loc["dim_count", col] = dm
                for target_col in df_count.columns:
                    for target_row in slice_matrix[target_col].dropna().index:
                        target_column = slice_matrix[target_col].dropna()
                        target_dim_item = target_column.loc[target_row]
                        df_rest_list = list(slice_matrix.columns)
                        df_rest_list.remove(target_col)
                        df_rest = slice_matrix.loc[:, df_rest_list]
                        x = 1
                        for item in df_rest.columns:
                            x = len(slice_matrix[item].dropna()) * x
                        df_item = pd.DataFrame(
                            index=range(0, x), columns=slice_matrix.columns
                        )
                        df_item[target_col] = target_dim_item
                        for rest_col in df_rest.columns:
                            rest_column = df_rest[rest_col].dropna()
                            rest_item_count = df_count[rest_col][0]
                            rest_multiplier = (
                                int(len(df_item.index) / rest_item_count) - 1
                            )
                            rest_column_insert = rest_column.copy()
                            for mlper in range(0, rest_multiplier):
                                rest_column_insert = rest_column_insert.append(
                                    rest_column
                                )
                            rest_column_insert.index = df_item.index
                            df_item[rest_col] = rest_column_insert

                        iter_matrix = pd.concat(
                            [iter_matrix, df_item], ignore_index=True
                        )

                iter_matrix = (
                    iter_matrix.drop_duplicates()
                    .dropna(how="any")
                    .reset_index(drop=True)
                )
                for row in iter_matrix.index:
                    slice_terms = tuple(iter_matrix.loc[row].values)
                    if isinstance(self.consolidation.index, pd.MultiIndex):
                        single_slice = self.consolidation.xs(
                            slice_terms, level=dims, drop_level=False
                        )
                    else:
                        for dim in dims:
                            for item_s in dim_values:
                                single_slice = self.consolidation.loc[item_s]
                    self.slice = pd.concat([self.slice, single_slice])
                    if isinstance(self.consolidation.index, pd.Index):
                        self.slice.drop_duplicates(inplace=True)
                self.slice.sort_index(inplace=True)
            # Column Search
            if col_search_dims != [] and col_search_values != []:
                col_slice = pd.DataFrame()
                col_slice_df = self.slice.copy()
                for dim in col_search_dims:
                    for val in col_search_values:
                        if isinstance(val, list):
                            for sub_val in val:
                                single_slice = col_slice_df[
                                    col_slice_df[dim] == sub_val
                                ]
                                col_slice = pd.concat([col_slice, single_slice])
                        else:
                            single_slice = col_slice_df[col_slice_df[dim] == val]
                            col_slice = pd.concat([col_slice, single_slice])
                    col_slice_df = col_slice
                self.slice = col_slice
            # Column axis slicing
            if col_range != None:
                self.slice = self.slice.loc[:, col_range[0] : col_range[1]]
            if col_list != None:
                self.slice = self.slice.loc[:, col_list]
            return self.slice


        if data_obj == "slice":
            df_slice = self.slice.copy()
            self.slice = pd.DataFrame()
            col_search_dims = []
            col_search_values = []
            col_search_dims = [
                dim
                for dim in dims
                if dim not in list(df_slice.index.names)
                and dim in list(df_slice.columns)
            ]
            dims_idx = [dims.index(col) for col in col_search_dims]
            col_search_values = [dim_values[idx] for idx in dims_idx]
            dims = [dim for dim in dims if dim not in col_search_dims]
            dim_values = [dim for dim in dim_values if dim not in col_search_values]
            if append_to == False:
                self.slice = pd.DataFrame()
            if dims == None or dims == [None] or dims == []:
                self.slice = df_slice.copy()
            else:
                slice_matrix = pd.DataFrame(columns=dims)
                if dim_values == None or dim_values == [None]:
                    dim_values = []
                    if isinstance(df_slice.index, pd.MultiIndex):
                        for item in df_slice.index.levels:
                            dim_values.append(list(item))
                    else:
                        dim_values = df_slice.index.values
                count_row = 0
                for item, dim_name in zip(dim_values, dims):
                    sub_count_row = 0
                    if isinstance(item, list):
                        for subitem in item:
                            slice_matrix.loc[sub_count_row, dim_name] = subitem
                            sub_count_row += 1
                    else:
                        slice_matrix.loc[count_row, dim_name] = item

                iter_matrix = pd.DataFrame()
                df_count = pd.DataFrame(
                    index=["dim_count"], columns=slice_matrix.columns
                )
                for col in slice_matrix.columns:
                    dm = len(slice_matrix[col].dropna())
                    df_count.loc["dim_count", col] = dm
                for target_col in df_count.columns:
                    for target_row in slice_matrix[target_col].dropna().index:
                        target_column = slice_matrix[target_col].dropna()
                        target_dim_item = target_column.loc[target_row]
                        df_rest_list = list(slice_matrix.columns)
                        df_rest_list.remove(target_col)
                        df_rest = slice_matrix.loc[:, df_rest_list]
                        x = 1
                        for item in df_rest.columns:
                            x = len(slice_matrix[item].dropna()) * x
                        df_item = pd.DataFrame(
                            index=range(0, x), columns=slice_matrix.columns
                        )
                        df_item[target_col] = target_dim_item
                        for rest_col in df_rest.columns:
                            rest_column = df_rest[rest_col].dropna()
                            rest_item_count = df_count[rest_col][0]
                            rest_multiplier = (
                                int(len(df_item.index) / rest_item_count) - 1
                            )
                            rest_column_insert = rest_column.copy()
                            for mlper in range(0, rest_multiplier):
                                rest_column_insert = rest_column_insert.append(
                                    rest_column
                                )
                            rest_column_insert.index = df_item.index
                            df_item[rest_col] = rest_column_insert

                        iter_matrix = pd.concat(
                            [iter_matrix, df_item], ignore_index=True
                        )

                iter_matrix = (
                    iter_matrix.drop_duplicates()
                    .dropna(how="any")
                    .reset_index(drop=True)
                )
                for row in iter_matrix.index:
                    slice_terms = tuple(iter_matrix.loc[row].values)
                    if isinstance(df_slice.index, pd.MultiIndex):
                        single_slice = df_slice.xs(
                            slice_terms, level=dims, drop_level=False
                        )
                    else:
                        for dim in dims:
                            for item_s in dim_values:
                                single_slice = self.data.loc[item_s]
                    self.slice = pd.concat([self.slice, single_slice])
                    if isinstance(self.slice.index, pd.Index):
                        self.slice.drop_duplicates(inplace=True)
                self.slice.sort_index(inplace=True)
            # Column Search
            if col_search_dims != [] and col_search_values != []:
                col_slice = pd.DataFrame()
                col_slice_df = self.slice.copy()
                for dim in col_search_dims:
                    for val in col_search_values:
                        if isinstance(val, list):
                            for sub_val in val:
                                single_slice = col_slice_df[
                                    col_slice_df[dim] == sub_val
                                ]
                                col_slice = pd.concat([col_slice, single_slice])
                        else:
                            single_slice = col_slice_df[col_slice_df[dim] == val]
                            col_slice = pd.concat([col_slice, single_slice])
                    col_slice_df = col_slice
                self.slice = col_slice
            # Column axis slicing
            if col_range != None:
                self.slice = self.slice.loc[:, col_range[0] : col_range[1]]
            if col_list != None:
                self.slice = self.slice.loc[:, col_list]
            return self.slice

    def column_slice(
        self,
        dims=None,
        dim_values=None,
        col_range=None,
        col_list=None,
        data_obj="data",
        append_to=False,
    ):

        dims = dims if isinstance(dims, list) else [dims]
        dim_values = dim_values if isinstance(dim_values, list) else [dim_values]
        if data_obj == "data":
            if append_to == False:
                self.slice = pd.DataFrame()
            if dims == None or dims == [None]:
                self.slice = self.data.copy()

    def time_slice(
        self,
        dim=None,
        start_dt=None,
        end_dt=None,
        idx_period=None,
        col_period=None,
        data_obj="data",
    ):
        if data_obj == "data":
            original_order = list(self.data.index.names)
            if dim != None:
                slice_order = original_order.copy()
                slice_order.remove(dim)
                slice_order.insert(0, dim)
                self.reorder_dimensions(slice_order, "data")
                self.slice = self.data.xs(slice(start_dt, end_dt), 0, dim, drop_level=False)
                self.reorder_dimensions(original_order, "data")
            else:
                self.slice = self.data.copy()
            if idx_period != None:
                self.slice = self.slice.resample(idx_period, 0, level=dim).sum()
            if col_period != None:
                self.slice = self.slice.resample(col_period, 1).sum()
            if isinstance(self.slice.index, pd.MultiIndex):
                self.reorder_dimensions(original_order, "slice")
            return self.slice
        elif data_obj == "slice":
            original_order = list(self.slice.index.names)
            if dim != None:
                slice_order = original_order.copy()
                slice_order.remove(dim)
                slice_order.insert(0, dim)
                self.reorder_dimensions(slice_order, "slice")
                self.slice = self.slice.xs(
                    slice(start_dt, end_dt), 0, dim, drop_level=False
                )
            if idx_period != None:
                self.slice = self.slice.resample(idx_period, 0, level=dim).sum()
            if col_period != None:
                self.slice = self.slice.resample(col_period, 1).sum()
            if isinstance(self.slice.index, pd.MultiIndex):
                self.reorder_dimensions(original_order, "slice")
            return self.slice

    def keyword_slice(self, keywords, dims=None, data_obj="data"):
        """
        Slice data based on string fragment.  Function will search the index & data and return any rows with the
        search string.

        :param keywords: String object of which to search.  I.e. 'Rent', 'LLC', 'Smith'
        :param dims: String or list object of specific dimensions to search. Blank will search all dimensions & data.
        :param data_obj: Which data object you want o effect.  Default: 'data'. Available - 'data', 'slice'
        :return: self.slice
        """
        if data_obj == "data":
            keywords = [keywords] if isinstance(keywords, str) else keywords
            self.slice = self.data.copy()
            original_index_order = self.slice.index.names
            if dims == None:
                dims = original_index_order.copy()
                dims.remove("Data_Block")
            self.slice.reset_index(dims, inplace=True)
            keyword_matrix = pd.DataFrame()
            keyword_result = pd.DataFrame()
            for col in self.slice.columns:
                for word in keywords:
                    try:
                        keyword_result = self.slice[
                            self.slice[col].str.contains(word, case=False, na=False)
                        ]
                        keyword_matrix = pd.concat([keyword_matrix, keyword_result])
                    except:
                        pass

        elif data_obj == "slice":
            keywords = [keywords] if isinstance(keywords, str) else keywords
            original_index_order = self.slice.index.names
            if dims == None:
                dims = original_index_order.copy()
                dims.remove("Data_Block")
            self.slice.reset_index(dims, inplace=True)
            keyword_matrix = pd.DataFrame()
            keyword_result = pd.DataFrame()
            for col in self.slice.columns:
                for word in keywords:
                    try:
                        keyword_result = self.slice[
                            self.slice[col].str.contains(word, case=False, na=False)
                        ]
                    except:
                        pass
                    keyword_matrix = pd.concat([keyword_matrix, keyword_result])

        self.slice = keyword_matrix
        self.slice.set_index(dims, append=True, inplace=True)
        self.reorder_dimensions(original_index_order, "slice")
        return self.slice

    def keyword_replace(self, target_words, replace_words, dims=None, data_obj="data"):
        """
        Replace a word or words in the index.  For example, if you wanted to replace ['Payroll', 'East'] with
        ['Salaries & Wages', 'Northeast'].

        :param target_words: String or list of word(s) to be replace. I.e. 'Software' or ['Payroll', 'East'].
        :param replace_words: String or list of word(s) to insert. I.e. '3rd Party Code' or ['Salaries & Wages', 'Northeast'].
        :param dims: String or list of specific dimensions to effect.
        :param data_obj: Which data object you want o effect.  Default: 'data'. Available - 'data', 'slice'
        :return: data_obj
        """
        if data_obj == "data":
            target_words = (
                [target_words] if isinstance(target_words, str) else target_words
            )
            replace_words = (
                [replace_words] if isinstance(replace_words, str) else replace_words
            )
            original_index_order = self.data.index.names
            if dims == None:
                dims = original_index_order.copy()
                dims.remove("Data_Block")
            self.data.reset_index(dims, inplace=True)
            for word, rword in zip(target_words, replace_words):
                for col in self.data.columns:
                    try:
                        self.data[col] = self.data[col].str.replace(word, rword)
                    except:
                        pass
            self.data.set_index(dims, append=True, inplace=True)
            self.reorder_dimensions(original_index_order, "data")
            return self.data
        elif data_obj == "slice":
            target_words = (
                [target_words] if isinstance(target_words, str) else target_words
            )
            replace_words = (
                [replace_words] if isinstance(replace_words, str) else replace_words
            )
            original_index_order = self.slice.index.names
            if dims == None:
                dims = original_index_order.copy()
                dims.remove("Data_Block")
            self.slice.reset_index(dims, inplace=True)
            for word, rword in zip(target_words, replace_words):
                for col in self.slice.columns:
                    try:
                        self.slice[col] = self.slice[col].str.replace(word, rword)
                    except:
                        pass
            self.slice.set_index(dims, append=True, inplace=True)
            self.reorder_dimensions(original_index_order, "slice")
            return self.slice

    def make_records_for_pivot(self, data_obj="data"):
        """
        Change the data object to a records format so it can be pasted into Excel in a
        pivot table friendly format.  Only works if the column index is a single level.

        :param data_obj: Which data object you want o effect. Default: 'data'.  Available - 'block', 'data', 'slice', 'consolidation',
         'function_result', 'variance'
        :return:
        """
        if data_obj == "data":
            df_records = self.data.stack()
            df_records.to_clipboard()
            return df_records
        elif data_obj == "slice":
            df_records = self.slice.stack()
            df_records.to_clipboard()
            return df_records
        elif data_obj == "block":
            df_records = self.block.stack()
            df_records.to_clipboard()
            return df_records
        elif data_obj == "variance":
            df_records = self.variance.stack()
            df_records.to_clipboard()
            return df_records
        elif data_obj == "consolidation":
            df_records = self.consolidation.stack()
            df_records.to_clipboard()
            return df_records
        elif data_obj == "function_result":
            df_records = self.consolidation.stack()
            df_records.to_clipboard()
            return df_records
        else:
            print("no object")

    def save_project(self, prj_name=None, path_name=None):
        """
        Saves a project as either a directory with pickle files or as a json object.

        :param prj_name: String object for name of the project.  Will make a directory in path_name if it doesn't exist.
        :param path_name: Path to directory or a full file path for json.  I.e. for json file
         'C:/Budgets/Budgets v1.json' or 'C:/Budgets/' for normal save
        :return: nothing
        """
        if path_name == None:
            path_name = os.getcwd()
        if path_name[-4:] == "json":
            container = pd.DataFrame(
                index=["data", "data_idx", "meta", "accounts", "accts_idx"],
                columns=["Load"],
            )
            container.loc["data", "Load"] = self.data.to_json(orient="index")
            container.loc["data_idx", "Load"] = self.data.index.to_frame().to_json(
                orient="index"
            )
            container.loc["meta", "Load"] = self.meta_block.to_json(orient="index")
            container.loc["accounts", "Load"] = self.accounts.to_json(orient="index")
            container.loc["accts_idx", "Load"] = self.accounts.index.to_frame().to_json(
                orient="index"
            )
            container.to_json(path_name, orient="index")
        else:
            if (os.getcwd()[-1] != "\\") or (os.getcwd()[-1] != "/"):
                path_name = path_name + "\\"
            if path_name == None:
                if not os.path.isdir(prj_name):
                    os.mkdir(prj_name)
            else:
                if not os.path.isdir(path_name + prj_name):
                    os.mkdir(path_name + prj_name)
            self.data.to_pickle(path_name + prj_name + "\\" + prj_name + " - data.pkl")
            self.meta_block.to_pickle(
                path_name + prj_name + "\\" + prj_name + " - meta_block.pkl"
            )
            self.accounts.to_pickle(
                path_name + prj_name + "\\" + prj_name + " - accounts.pkl"
            )
        # TODO: Add and test compression. SAVE & LOAD

    def load_project(self, path_name):
        """
        Imports the project by filling the 'data', 'meta_block' and 'accounts' data objects.

        :param path_name: Path to directory for normal save or to the json file. I.e. for json file
         'C:/Budgets/Budgets v1.json' or 'C:/Budgets/' for normal load.
        :return: project
        """
        if path_name[-4:] == "json":
            load_container = pd.read_json(path_name, orient="index")
            self.data = pd.read_json(load_container.loc["Load", "data"], orient="index")
            data_idx = pd.read_json(
                load_container.loc["Load", "data_idx"], orient="index"
            )
            self.data.index = pd.MultiIndex.from_frame(data_idx)
            self.meta_block = pd.read_json(
                load_container.loc["Load", "meta"], orient="index", convert_axes=False
            )
            self.meta_block.index = self.meta_block.index.astype("int")
            self.accounts = pd.read_json(
                load_container.loc["Load", "accounts"], orient="index"
            )
            accounts_idx = pd.read_json(
                load_container.loc["Load", "accts_idx"], orient="index"
            )
            self.data.index = pd.MultiIndex.from_frame(data_idx)
        else:
            if (os.getcwd()[-1] != "\\") or (os.getcwd()[-1] != "/"):
                path_name = path_name + "\\"
            file_list = pd.Series(os.listdir(path_name))
            f_name = file_list[file_list.str.contains("- data.pkl")].iloc[0]
            data_df = pd.read_pickle(path_name + f_name)
            self.data = data_df.copy()
            f_name = file_list[file_list.str.contains("- meta_block.pkl")].iloc[0]
            meta_df = pd.read_pickle(path_name + f_name)
            self.meta_block = meta_df.copy()
            f_name = file_list[file_list.str.contains("- accounts.pkl")].iloc[0]
            accounts_df = pd.read_pickle(path_name + f_name)
            self.accounts = accounts_df

    # TODO: Make big data stable

    def slice_to_project(self):
        x = 1

    def make_pivot_table(
        self,
        value_col,
        index_names,
        col_names,
        data_obj="data",
        function="sum",
        totals=True,
        total_names="Total",
    ):
        if data_obj == "data":
            df = self.data.copy()
            df.reset_index(inplace=True)
            self.function_result= pd.pivot_table(
                df,
                value_col,
                index_names,
                col_names,
                margins=totals,
                margins_name=total_names,
            )
            return self.function_result
        if data_obj == "slice":
            df = self.slice.copy()
            df.reset_index(inplace=True)
            self.function_result= pd.pivot_table(
                df,
                value_col,
                index_names,
                col_names,
                aggfunc=function,
                margins=True,
                margins_name=total_names,
            )
            return self.function_result

    def variance_analysis(self, dim_name, dim1, dim2, data_obj="data"):
        """
        Returns and difference and percent difference analysis based on two items with in a dimension. I.e.
        x.variance_analysis('Type', 'Actual', 'Budget') will produce the 'variance' data object with amount and
        percent differences.

        :param dim_name: String object of the name of dimension.  I.e. 'Type' or 'Region'.
        :param dim1: String object of the name of dimension item for first part of calculation.  I.e. 'Actual' or 'North'.
        :param dim2: String object of the name of dimension item for comparison.  I.e. 'Budget' or 'South'.
        :param data_obj: Which data object you want o effect.  Default: 'data'. Available - 'data', 'slice'
        :return: self.variance
        """
        if data_obj == "data":
            df1 = self.data.xs(dim1, level=dim_name)
            df1 = df1.droplevel("Data_Block")
            df2 = self.data.xs(dim2, level=dim_name)
            df2 = df2.droplevel("Data_Block")
        if data_obj == "slice":
            df1 = self.slice.xs(dim1, level=dim_name)
            df1 = df1.droplevel("Data_Block")
            df2 = self.slice.xs(dim2, level=dim_name)
            df2 = df2.droplevel("Data_Block")
        amt_var = df1 - df2
        pct_var = df1 / df2 - 1
        self.variance = pd.concat([amt_var, pct_var], keys=["Amount", "Percent"])
        return self.variance

    # TODO: refine - percent?

    def multiply_dim(
        self, dim_name, dim_vals=None, calc_name="New Item", data_obj="slice"
    ):
        """
        Multiply two or more dimension items.  For example, multiply units x price x discount with
        x.multiply_dim('Basis', ['Units', 'Price', 'Discount']
        :param dim_name: String object of dimension name.  I.e. 'Basis'.
        :param dim_vals: List object of dimension items to multiply.  I.e. ['Units', 'Price', 'Discount']
        :param calc_name: String object of new dimension item name.  I.e. 'Total_Revenue'
        :param data_obj: Which data object you want o effect.  Default: 'slice'. Available - 'data', 'slice'
        :return: self.function_result
        """
        if data_obj == "data":
            df = pd.DataFrame()
            if dim_vals == None:
                dim_vals = set(list(self.data.index.get_level_values(dim_name)))
            for item in dim_vals:
                df1 = self.data.xs(item, level=dim_name)
                df1 = df1.droplevel("Data_Block")
                df1_idx = df1.index
                df1.reset_index(drop=True, inplace=True)
                if df.empty:
                    df = df1
                else:
                    df = df * df1
            df.index = df1_idx
        if data_obj == "slice":
            df = pd.DataFrame()
            if dim_vals == None:
                dim_vals = set(list(self.slice.index.get_level_values(dim_name)))
            for item in dim_vals:
                df1 = self.slice.xs(item, level=dim_name)
                df1 = df1.droplevel("Data_Block")
                df1_idx = df1.index
                df1.reset_index(drop=True, inplace=True)
                if df.empty:
                    df = df1
                else:
                    df = df * df1
            df.index = df1_idx
        self.function_result= df
        if not isinstance(dim_name, list):
            dim_name = [dim_name]
        if not isinstance(calc_name, list):
            calc_name = [calc_name]
        self.add_dimensions(dim_name, calc_name, data_obj="function_result")
        self.import_xl(self.function_result.copy())
        return self.function_result

    def sum_dim(self, dim_name, dim_vals=None, calc_name="New Item", data_obj="slice"):
        """
        Add two or more dimension items.  For example, 'Rent' + 'Office Supplies' with
        FPA_OBJECT.sum_dim('Line Item', ['Rent', 'Office Supplies']

        :param dim_name: String object of dimension name.  I.e. 'Line Item'.
        :param dim_vals: List object of dimension items to multiply.  I.e. ['Rent', 'Office Supplies']
        :param calc_name: String object of new dimension item name.  I.e. 'Office Expense'
        :param data_obj: Which data object you want o effect.  Default: 'slice'. Available - 'data', 'slice'
        :return: self.function_result
        """
        if data_obj == "data":
            df = pd.DataFrame()
            if isinstance(dim_vals, str):
                dim_vals = [dim_vals]
            if dim_vals == None:
                dim_vals = set(list(self.data.index.get_level_values(dim_name)))
            for item in dim_vals:
                df1 = self.data.xs(item, level=dim_name)
                df1 = df1.droplevel("Data_Block")
                df1_idx = df1.index
                df1.reset_index(drop=True, inplace=True)
                if df.empty:
                    df = df1
                else:
                    df = df + df1
            df.index = df1_idx
        if data_obj == "slice":
            df = pd.DataFrame()
            if isinstance(dim_vals, str):
                dim_vals = [dim_vals]
            if dim_vals == None:
                dim_vals = set(list(self.slice.index.get_level_values(dim_name)))
            for item in dim_vals:
                df1 = self.slice.xs(item, level=dim_name)
                df1 = df1.droplevel("Data_Block")
                df1_idx = df1.index
                df1.reset_index(drop=True, inplace=True)
                if df.empty:
                    df = df1
                else:
                    df = df + df1
            df.index = df1_idx
        self.function_result= df
        if not isinstance(dim_name, list):
            dim_name = [dim_name]
        if not isinstance(calc_name, list):
            calc_name = [calc_name]
        self.add_dimensions(dim_name, calc_name, data_obj="function_result")
        self.import_xl(self.function_result.copy())
        return self.function_result

    def subtract_dim(
        self, dim_name, dim_vals=None, calc_name="New Item", data_obj="slice"
    ):
        """
        Subtract two or more dimension items.  For example, 'Total_Revenue' - 'COGS' with
        FPA_OBJECT.subtract_dim('IS_Category', ['Total_Revenue', 'COGS']

        :param dim_name: String object of dimension name.  I.e. 'IS_Category'.
        :param dim_vals: List object of dimension items to multiply.  I.e. ['Total_Revenue', 'COGS']
        :param calc_name: String object of new dimension item name.  I.e. 'Gross Profit'
        :param data_obj: Which data object you want o effect.  Default: 'slice'. Available - 'data', 'slice'
        :return: self.function_result
        """
        if data_obj == "data":
            df = pd.DataFrame()
            if isinstance(dim_vals, str):
                dim_vals = [dim_vals]
            if dim_vals == None:
                dim_vals = set(list(self.data.index.get_level_values(dim_name)))
            for item in dim_vals:
                df1 = self.data.xs(item, level=dim_name)
                df1 = df1.droplevel("Data_Block")
                df1_idx = df1.index
                df1.reset_index(drop=True, inplace=True)
                if df.empty:
                    df = df1
                else:
                    df = df - df1
            df.index = df1_idx
        if data_obj == "slice":
            df = pd.DataFrame()
            if isinstance(dim_vals, str):
                dim_vals = [dim_vals]
            if dim_vals == None:
                dim_vals = set(list(self.slice.index.get_level_values(dim_name)))
            for item in dim_vals:
                df1 = self.slice.xs(item, level=dim_name)
                df1 = df1.droplevel("Data_Block")
                df1_idx = df1.index
                df1.reset_index(drop=True, inplace=True)
                if df.empty:
                    df = df1
                else:
                    df = df - df1
            df.index = df1_idx
        self.function_result= df
        if not isinstance(dim_name, list):
            dim_name = [dim_name]
        if not isinstance(calc_name, list):
            calc_name = [calc_name]
        self.add_dimensions(dim_name, calc_name, data_obj="function_result")
        self.import_xl(self.function_result.copy())
        return self.function_result

    def remove_duplicates(self, based_on=None, keep_item="last", data_obj="data"):
        """
        Deletes repetitive records based on the index.  Can be filtered down by dimensions.

        :param based_on: String or list object of dimensions to be used.  If blank it will all dimensions.
        :param keep_item: String object of which records to keep. I.e. either 'last' or 'first'.
        :param data_obj: Which data object you want o effect.  Default: 'data'. Available - 'data', 'block', 'slice'
        :return: data object without duplicate records
        """
        if data_obj == "data":
            lx = list(self.data.index.names)
            index_order = list(self.data.index.names)
            lx.remove("Data_Block")
            self.data.reset_index(level=lx, inplace=True)
            if based_on == None:
                based_on = list(self.data.columns)
            if isinstance(based_on, str):
                based_on = [based_on]
            self.data.drop_duplicates(subset=based_on, keep=keep_item, inplace=True)
            self.data.set_index(lx, append=True, inplace=True)
            self.reorder_dimensions(index_order, "data")
            return self.data
        elif data_obj == "block":
            lx = list(self.block.index.names)
            index_order = list(self.block.index.names)
            lx.remove("Data_Block")
            self.block.reset_index(level=lx, inplace=True)
            if based_on == None:
                based_on = list(self.block.columns)
            if isinstance(based_on, str):
                based_on = [based_on]
            self.block.drop_duplicates(subset=based_on, keep=keep_item, inplace=True)
            self.block.set_index(lx, append=True, inplace=True)
            self.reorder_dimensions(index_order, "block")
            return self.block
        elif data_obj == "slice":
            lx = list(self.slice.index.names)
            index_order = list(self.slice.index.names)
            lx.remove("Data_Block")
            self.slice.reset_index(level=lx, inplace=True)
            if based_on == None:
                based_on = list(self.slice.columns)
            if isinstance(based_on, str):
                based_on = [based_on]
            self.slice.drop_duplicates(subset=based_on, keep=keep_item, inplace=True)
            self.slice.set_index(lx, append=True, inplace=True)
            self.reorder_dimensions(index_order, "block")
            return self.slice

    def get_duplicates(self, based_on=None, data_obj="data"):
        """
        Retrieves repetitive records based on the index.  Can be filtered down by dimensions.

        :param based_on: String or list object of dimensions to be used.  If blank it will all dimensions.
        :param data_obj: Which data object you want o effect.  Default: 'data'. Available - 'data'
        :return: self.function_result
        """
        if data_obj == "data":
            lx = list(self.data.index.names)
            index_order = list(self.data.index.names)
            lx.remove("Data_Block")
            self.function_result= self.data.copy()
            self.function_result.reset_index(level=lx, inplace=True)
            if based_on == None:
                based_on = list(self.function_result.columns)
            if isinstance(based_on, str):
                based_on = [based_on]
            try:
                self.function_result= pd.concat(
                    z for _, z in self.function_result.groupby(based_on) if len(z) > 1
                )
            except ValueError as e:
                if str(e) == "No objects to concatenate":
                    print("No duplicates.")
                    pass
            self.function_result.set_index(lx, append=True, inplace=True)
            self.reorder_dimensions(index_order, "function_result")
            return self.function_result

    def _align_indicies(self):
        x = self.block.index.names  # TODO: move this to update_custom_xl
        y = self.data.index.names
        z = [item in x for item in y]
        w = zip(y, z)
        for item in w:
            if item[1] == False:
                self.add_dimensions([item[0]], ["na"], data_obj='block')
        z = [item in y for item in x]
        w = zip(x, z)
        for item in w:
            if item[1] == False:
                self.add_dimensions([item[0]], ["na"], 1, "data")